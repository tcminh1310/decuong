from django.shortcuts import render
from django.http import HttpResponse
from django.views import View
import openpyxl as xl
import xml.etree.ElementTree as ET
import uuid
from . import dttype as dt
import requests
# Create your views here.
class index(View):
    def get(self, request):
        return render(request, 'myapp/index.html')

class upload(View):
    def get(self, request):
        return render(request, 'myapp/upload.html')
    
    def post(self, request):
        if request.FILES.get('excel_file'):
            excel_file = request.FILES.get('excel_file')
            # print((excel_file))
            # tieude = mail.cleaned_data['title']
            # noidung = mail.cleaned_data['content']
            # email = mail.cleaned_data['email']
            # cc = mail.cleaned_data['email']
            # _file = request.FILES['_file']
            wb = xl.load_workbook(excel_file)
            sh = wb['Sheet1']
            m_row = sh.max_row
            last_line = 0
            data = {'Patient':{},'Encounter':{},'Observation':[]}
            for i in range(1,m_row+1):
                cell = sh.cell(row=i, column=4)
                if(cell.value):
                    tag = sh.cell(row=i, column=2)
                    if tag.value == 'HO_TEN':
                        data['Patient']['name'] = cell.value
                    elif tag.value == 'NGAY_SINH':
                        data['Patient']['birthDate'] = cell.value
                    elif tag.value == 'GIOI_TINH':
                        data['Patient']['gender'] = cell.value
                    elif tag.value == 'DIA_CHI':
                        data['Patient']['address'] = [{'address':cell.value}]
                    elif tag.value == 'NGAY_VAO':
                        data['Encounter']['period'] = {'start':cell.value}
                    elif tag.value == 'NGAY_RA':
                        if data['Encounter'].get('period'):
                            data['Encounter']['period']['stop'] = cell.value
                        else:
                            data['Encounter']['period'] = {'stop':cell.value}
                    elif tag.value == 'MA_LOAI_KCB':
                        data['Encounter']['class'] = cell.value
                    elif tag.value == 'MA_KHOA':
                        data['Encounter']['location'] = cell.value
                    elif tag.value == 'SO_NGAY_DTRI':
                        data['Encounter']['length'] = cell.value                        
                    elif not tag.value:
                        tag_2 = sh.cell(row=i, column=3)
                        tag_2_content = tag_2.value.split('.')
                        if tag_2_content[0] == 'Patient':
                            if data['Patient'][tag_2_content[1]]:
                                data['Patient'][tag_2_content[1]].append({tag_2_content[1]:cell.value})
                                if len(tag_2_content) > 2:
                                    for i in range(2, len(tag_2_content)):
                                        data['Patient'][tag_2_content[1]][-1][tag_2_content[i].split('=')[0]] = tag_2_content[i].split('=')[1]
                        elif tag_2_content[0] == 'Encounter':
                            if data['Encounter'][tag_2_content[1]]:
                                data['Encounter'][tag_2_content[1]].append({tag_2_content[1]:cell.value})
                                if len(tag_2_content) > 2:
                                    for i in range(2, len(tag_2_content)):
                                        data['Encounter'][tag_2_content[1]][-1][tag_2_content[i].split('=')[0]] = tag_2_content[i].split('=')[1]
                        elif tag_2_content[0] == 'Observation':
                            _observation = {}
                            for i in range(1, len(tag_2_content)):
                                _observation[tag_2_content[i].split('=')[0]] = tag_2_content[i].split('=')[1]
                            _observation['valueQuantity'] = cell.value
                            data['Observation'].append(_observation)

            # print(data)                            
            id_system = ""
            data['Patient']['identifier'] = '12345'
            root = ET.Element('Patient')
            tree = ET.ElementTree(root)
            root.set("xmlns","http://hl7.org/fhir")
            if data['Patient'].values():
                identifier = ET.SubElement(root, 'identifier')
                dt.identifier_type(identifier, 'urn:trinhcongminh', '12345', 'usual',{'codes': [{'system': 'http://terminology.hl7.org/CodeSystem/v2-0203', 'code': 'MR'}]})
                if data['Patient'].get('name'):
                    name = ET.SubElement(root, 'name')
                    dt.name_type(name, data['Patient']['name'])
                if data['Patient'].get('gender'):
                    gender = ET.SubElement(root, 'gender')
                    if data['Patient']['gender'] == 'Nam':
                        code = 'male'
                    elif data['Patient']['gender'] == 'Nữ':
                        code = 'female'
                    gender.set('value', code)
                if data['Patient'].get('birthDate'):
                    birthDate = ET.SubElement(root, 'birthDate')
                    birthDate.set('value', data['Patient']['birthDate'])
                if data['Patient'].get('address'):
                    for value in data['Patient']['address']:
                        address = ET.SubElement(root,'address')
                        dt.address_type(address, value.get('address'),value.get('postalCode'),value.get('country'),value.get('use'),value.get('type'))
                if data['Patient'].get('contact'):
                    contact = ET.SubElement(root, 'contact')
                

            put_req = None
            post_req = None
            encounter_id = None
            # print(text)
            get_req = requests.get("http://hapi.fhir.org/baseR4/Patient?identifier=urn:trinhcongminh|" +'12345', headers={'Content-type': 'application/xml'})
            if get_req.status_code == 200 and 'entry' in get_req.content.decode('utf-8'):
                print(get_req.status_code)
                get_root = ET.fromstring(get_req.content.decode('utf-8'))
                ns = {'d':"http://hl7.org/fhir"}
                entry = get_root.find('d:entry', ns)
                resource = entry.find('d:resource', ns)
                patient_resource = resource.find('d:Patient', ns)
                id_resource = patient_resource.find('d:id', ns)
                patient_id = id_resource.attrib['value']
                # print(_id)
                root.insert(0, ET.Element('id'))
                res_id = root.find('id')
                res_id.set('value', patient_id)
                text = ET.tostring(root, encoding="us-ascii", method="xml", xml_declaration=None, default_namespace=None, short_empty_elements=True)
                put_req = requests.put("http://hapi.fhir.org/baseR4/Patient/"+patient_id, headers={'Content-type': 'application/xml'}, data=text.decode('utf-8'))
                print(ET.tostring(root, encoding="us-ascii", method="xml", xml_declaration=None, default_namespace=None, short_empty_elements=True))
            # print(type(text.decode('utf-8')))
            # print(text.decode('utf-8'))
            else: 
                text = ET.tostring(root, encoding="us-ascii", method="xml", xml_declaration=None, default_namespace=None, short_empty_elements=True)
                post_req = requests.post("http://hapi.fhir.org/baseR4/Patient/", headers={'Content-type': 'application/xml'}, data=text.decode('utf-8'))
                if post_req.status_code == 201:
                    get_root = ET.fromstring(get_req.content.decode('utf-8'))
                    ns = {'d':"http://hl7.org/fhir"}
                    id_resource = get_root.find('d:id', ns)
                    patient_id = id_resource.attrib['value']                         
            # print(x.status_code)
            # print(x.content)
            # print(x.content.decode('utf-8'))
            if (put_req and put_req.status_code == 200) or (post_req and post_req.status_code == 201):
                if data['Encounter']:
                    root = ET.Element('Encounter')
                    tree = ET.ElementTree(root)
                    root.set("xmlns","http://hl7.org/fhir")
                    if data['Encounter'].values():
                        ids = ET.SubElement(root, 'id')
                        ids.set('value', '1')
                        # identifier = ET.SubElement(root, 'identifier')
                        # dttype.identifier_type(identifier, )
                        if not data['Encounter'].get('status'):
                            status = ET.SubElement(root, 'status')
                            status.set('value', 'in-progress')
                        else:
                            status = ET.SubElement(root, 'status')
                            status.set('value', data['Encounter']['status'])
                        if data['Encounter'].get('class'):
                            _class = ET.SubElement(root, 'class')
                            dt.coding_type(_class, 'http://terminology.hl7.org/CodeSystem/v3-ActCode', data['Encounter']['class'])
                        if data['Encounter'].get('serviceType'):
                            serviceType = ET.SubElement(root, 'serviceType')
                        subject = ET.SubElement(root, 'subject')
                        dt.reference_type(subject, 'Patient/'+ patient_id, 'Patient', display=data['Patient']['name'])
                        if data['Encounter'].get('period'):
                            period = ET.SubElement(root, 'period')
                            dt.period_type(period, data['Encounter']['period'].get('start'), data['Encounter']['period'].get('stop'))
                        if data['Encounter'].get('length'):
                            length = ET.SubElement(root, 'length')
                            dt.duration_type(length, data['Encounter']['length'], 'days', 'http://unitsofmeasure.org', 'd')    
                        # if data['Encounter'].get('location'):
                        #     location = ET.SubElement(root, 'location')
                        if data['Encounter'].get('serviceProvider'):
                            serviceProvider = ET.SubElement(root, 'serviceProvider')
                    text = ET.tostring(root, encoding="us-ascii", method="xml", xml_declaration=None, default_namespace=None, short_empty_elements=True)
                    post_req = requests.post("http://hapi.fhir.org/baseR4/Encounter/", headers={'Content-type': 'application/xml'}, data=text.decode('utf-8'))
                    # print(post_req.status_code)
                    if post_req.status_code == 201:
                        get_root = ET.fromstring(post_req.content.decode('utf-8'))
                        ns = {'d':"http://hl7.org/fhir"}
                        id_resource = get_root.find('d:id', ns)
                        encounter_id = id_resource.attrib['value']  
                        print(encounter_id)
                if data['Observation']:
                    for i in range(len(data['Observation'])):
                        root = ET.Element('Observation')
                        tree = ET.ElementTree(root)
                        root.set('xmlns', 'http://hl7.org/fhir')
                        ids = ET.SubElement(root, 'id')
                        ids.set('value', '{}'.format(i))
                        if data['Observation'][i].get('status'):
                            status = ET.SubElement(root, 'status')
                            status.set('value', data['Observation'][i]['status'])
                        if data['Observation'][i].get('category'):
                            category = ET.SubElement(root, 'category')
                            coding = ET.SubElement(category, 'coding')
                            dt.coding_type(coding, 'http://terminology.hl7.org/CodeSystem/observation-category', data['Observation'][i]['category'])
                        if data['Observation'][i].get('code'):
                            code = ET.SubElement(root, 'code')
                            if patient_id:
                                subject = ET.SubElement(root, 'subject')
                                dt.reference_type(subject, 'Patient/' + patient_id, 'Patient', display=data['Patient']['name'])
                            if encounter_id:
                                encounter = ET.SubElement(root, 'encounter')
                                dt.reference_type(encounter, 'Encounter/'+encounter_id, 'Encounter')
                            if data['Observation'][i]['code'] == '8867-4':
                                coding = ET.SubElement(code, 'coding')
                                dt.coding_type(coding, 'http://loinc.org', '8867-4', display='Heart rate')
                                valueQuantity = ET.SubElement(root, 'valueQuantity')
                                dt.quantity_type(valueQuantity, data['Observation'][i]['valueQuantity'], 'beats/minute', 'http://unitsofmeasure.org', '{Beats}/min')
                            elif data['Observation'][i]['code'] == '8310-5':
                                coding = ET.SubElement(code, 'coding')
                                dt.coding_type(coding, 'http://loinc.org', '8867-4', display='Body temperature')
                                valueQuantity = ET.SubElement(root, 'valueQuantity')
                                dt.quantity_type(valueQuantity, data['Observation'][i]['valueQuantity'], 'Cel', 'http://unitsofmeasure.org', 'Cel')
                            elif data['Observation'][i]['code'] == '8480-6':
                                coding = ET.SubElement(code, 'coding')
                                dt.coding_type(coding, 'http://loinc.org', '8480-6', display='Systolic blood pressure')
                                valueQuantity = ET.SubElement(root, 'valueQuantity')
                                dt.quantity_type(valueQuantity, data['Observation'][i]['valueQuantity'], 'mmHg', 'http://unitsofmeasure.org', 'mm[Hg]')
                            elif data['Observation'][i]['code'] == '8462-4':
                                coding = ET.SubElement(code, 'coding')
                                dt.coding_type(coding, 'http://loinc.org', '8462-4', display='Diastolic blood pressure')
                                valueQuantity = ET.SubElement(root, 'valueQuantity')
                                dt.quantity_type(valueQuantity, data['Observation'][i]['valueQuantity'], 'mmHg', 'http://unitsofmeasure.org', 'mm[Hg]')
                            elif data['Observation'][i]['code'] == '9279-1':
                                coding = ET.SubElement(code, 'coding')
                                dt.coding_type(coding, 'http://loinc.org', '9279-1', display='Respiratory rate')
                                valueQuantity = ET.SubElement(root, 'valueQuantity')
                                dt.quantity_type(valueQuantity, data['Observation'][i]['valueQuantity'], 'breaths/minute', 'http://unitsofmeasure.org', '{Breaths}/min')
                            elif data['Observation'][i]['code'] == '29463-7':
                                coding = ET.SubElement(code, 'coding')
                                dt.coding_type(coding, 'http://loinc.org', '29463-7', display='Body weight')
                                valueQuantity = ET.SubElement(root, 'valueQuantity')
                                dt.quantity_type(valueQuantity, data['Observation'][i]['valueQuantity'], 'kg', 'http://unitsofmeasure.org', 'kg')
                        text = ET.tostring(root, encoding="us-ascii", method="xml", xml_declaration=None, default_namespace=None, short_empty_elements=True)
                        # print(text)
                        post_req = requests.post("http://hapi.fhir.org/baseR4/Observation/", headers={'Content-type': 'application/xml'}, data=text.decode('utf-8'))
                        
                        # if post_req.status_code == 201:
                        #     get_root = ET.fromstring(get_req.content.decode('utf-8'))
                        #     ns = {'d':"http://hl7.org/fhir"}
                        #     id_resource = get_root.find('d:id', ns)
                            # encounter_id = id_resource.attrib['value']         
                return render(request, 'myapp/display.html', {'message': 'Upload successful', 'patient': data['Patient']})
            else: return render(request, 'myapp/index.html',{'message': 'Failed to create resource, please check your file!'})
        else: return render(request, 'myapp/index.html',{'message':'Please upload your file!'})


class search(View):
    def post(self, request):
        if request.POST:
            x = requests.get("http://hapi.fhir.org/baseR4/Patient?identifier=urn:trinhcongminh|" + request.POST['identifier'], headers={'Content-type': 'application/xml'})
            if x.status_code == 200 and 'entry' in x.content.decode('utf-8'):
                return render(request, 'myapp/index.html', {'message': 'Patient data existed'})
            else: 
                return render(request, 'myapp/index.html', {'message': 'Patient not found in database'})
        else:
            return render(request, 'myapp/index.html', {'message':'Please enter an identifier'})

class register(View):
    def get(self, request):
        return render(request, 'myapp/register.html')
    def post(self, request):
        if request.POST:
            patient = {}
            patient['name'] = request.POST['ho_ten']
            patient['gender'] = request.POST['gioi_tinh']
            patient['birthDate'] = request.POST['ngay_sinh']
            patient['address'] = [{'address': request.POST['dia_chi'], 'use': 'home'}, {'address': request.POST['noi_lam_viec'], 'use': 'work'}]
            patient['identifier'] =str(uuid.uuid4())
            print(type(patient['identifier']))
            print(patient)
            id_system = "urn:trinhcongminh"
            root = ET.Element('Patient')
            tree = ET.ElementTree(root)
            root.set("xmlns","http://hl7.org/fhir")
            if patient.values():
                if patient.get('identifier'):
                    identifier = ET.SubElement(root, 'identifier')
                    dt.identifier_type(identifier, id_system, patient['identifier'], 'usual',{'codes': [{'system': 'http://terminology.hl7.org/CodeSystem/v2-0203', 'code': 'MR'}]})
                if patient.get('name'):
                    # for value in patient.get('name'):
                    name = ET.SubElement(root, 'name')
                    dt.name_type(name, patient['name'])
                if patient.get('gender'):
                    gender = ET.SubElement(root, 'gender')
                    if patient['gender'] == 'Nam':
                        code = 'male'
                    elif patient['gender'] == 'Nữ':
                        code = 'female'
                    gender.set('value', code)
                if patient.get('birthDate'):
                    birthDate = ET.SubElement(root, 'birthDate')
                    birthDate.set('value', patient['birthDate'])
                if patient.get('address'):
                    for value in patient['address']:
                        address = ET.SubElement(root,'address')
                        dt.address_type(address, value.get('address'),value.get('postalCode'),value.get('country'),value.get('use'),value.get('type'))
                if patient.get('contact'):
                    contact = ET.SubElement(root, 'contact')
            text = ET.tostring(root, encoding="us-ascii", method="xml", xml_declaration=None, default_namespace=None, short_empty_elements=True)
            # print(type(text.decode('utf-8')))
            # print(text.decode('utf-8'))
            x = requests.post("http://hapi.fhir.org/baseR4/Patient/", headers={'Content-type': 'application/xml'}, data=text.decode('utf-8'))
            return render(request, 'myapp/display.html', {'patient':patient})
        else: return HttpResponse("Please enter your information")
    